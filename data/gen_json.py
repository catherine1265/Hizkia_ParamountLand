# -*- coding: utf-8 -*-
"""
Generator database terpusat proyek.json dari:
  - daftar-proyek.xlsx (sheet Proyek + Tipe Unit)
  - folder foto-proyek/[slug]/ (cover, foto-*, denah-*)
  - folder brosur/ (brosur-[slug].pdf)
Menghasilkan JSON kaya: harga display, range, statistik unit, label foto,
fasilitas terstruktur, agregat/statistik global. Idempotent — jalankan ulang
kapan saja setelah data berubah.
"""
import openpyxl, json, os, re

DATA = r"c:\Users\Sharon\paramount-website\data"
XLSX = os.path.join(DATA, "daftar-proyek.xlsx")
FOTO_BASE = os.path.join(DATA, "foto-proyek")
BROSUR_BASE = os.path.join(DATA, "brosur")
OUT = os.path.join(DATA, "proyek.json")

# ---------------- info tambahan hasil bacaan brosur ----------------
INFO_TAMBAHAN = {}
KONTEKS_KOTA = {}
_info_path = os.path.join(DATA, "brosur-info-bersih.json")
if os.path.isfile(_info_path):
    import json as _json
    _d = _json.load(open(_info_path, encoding="utf-8"))
    INFO_TAMBAHAN = _d.get("info", {})
    KONTEKS_KOTA = _d.get("konteks_kota", {})

# ---------------- util ----------------
def rupiah_singkat(n):
    """2500000000 -> 'Rp 2,5 M' ; 750000000 -> 'Rp 750 Jt'"""
    if not isinstance(n, (int, float)):
        return None
    if n >= 1_000_000_000:
        v = n / 1_000_000_000
        s = f"{v:.2f}".rstrip("0").rstrip(".").replace(".", ",")
        return f"Rp {s} M"
    if n >= 1_000_000:
        v = n / 1_000_000
        s = f"{v:.0f}"
        return f"Rp {s} Jt"
    return f"Rp {n:,.0f}".replace(",", ".")

def natkey(s):
    return [int(x) if x.isdigit() else x for x in re.split(r"(\d+)", s)]

# fasilitas/POI terstruktur dari deskripsi & keunggulan (keyword matching)
POI_KEYWORDS = {
    "Sekolah/Universitas": ["sekolah", "universitas", "univ ", "tarakanita", "stella maris", "matana", "pendidikan", "international school"],
    "Rumah Sakit": ["rumah sakit", "bethsaida", "rs ", "hospital", "kesehatan"],
    "Pusat Perbelanjaan": ["mal", "mall", "summarecon", "pusat perbelanjaan", "pasar modern", "paramount modern market", "shopping", "bez", "g-town", "gtown"],
    "Akses Tol": ["tol", "gerbang tol", "toll"],
    "Area Komersial": ["komersial", "ruko", "kuliner", "restoran", "kafe", "boulevard"],
    "Danau/Area Hijau": ["danau", "lakeside", "lake", "hijau", "verde", "taman", "green"],
}
FASILITAS_KEYWORDS = {
    "Clubhouse": ["clubhouse", "club house", "community club"],
    "Kolam Renang": ["kolam renang", "swimming"],
    "Keamanan 24 Jam": ["24 jam", "one gate", "gated", "keamanan", "security"],
    "Jogging Track": ["jogging"],
    "Gym": ["gym", "gymnasium"],
    "Carport": ["carport"],
}

def extract_tags(text, mapping):
    if not text:
        return []
    low = text.lower()
    out = []
    for label, kws in mapping.items():
        if any(k in low for k in kws):
            out.append(label)
    return out

# ---------------- baca excel ----------------
wb = openpyxl.load_workbook(XLSX)
ws_proj, ws_tipe = wb["Proyek"], wb["Tipe Unit"]

def headers(ws):
    return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

def rows(ws, hdr, start=4):
    out = []
    for r in range(start, ws.max_row + 1):
        if not ws.cell(r, 1).value:
            continue
        out.append({name: ws.cell(r, c).value for c, name in enumerate(hdr, 1)})
    return out

proj_rows = rows(ws_proj, headers(ws_proj))
tipe_rows = rows(ws_tipe, headers(ws_tipe))

# group tipe by slug
tipe_by_slug = {}
for t in tipe_rows:
    tipe_by_slug.setdefault(t["slug_proyek"], []).append(t)

# ---------------- scan aset ----------------
def label_foto(fname):
    low = fname.lower()
    if low.startswith("cover."): return "cover"
    if low.startswith("denah-"): return "denah"
    if low.startswith("foto-"):  return "galeri"
    if low.startswith("brosur-"): return "brosur"
    return "lain"

def scan_foto(slug):
    d = os.path.join(FOTO_BASE, slug)
    cover, galeri, denah, brosur = None, [], [], []
    if os.path.isdir(d):
        for f in sorted(os.listdir(d), key=natkey):
            if not os.path.isfile(os.path.join(d, f)):
                continue
            rel = f"foto-proyek/{slug}/{f}"
            lab = label_foto(f)
            if lab == "cover": cover = rel
            elif lab == "galeri": galeri.append(rel)
            elif lab == "denah": denah.append(rel)
            elif lab == "brosur": brosur.append(rel)
    if cover is None:
        cover = (galeri or brosur or denah or [None])[0]
    return cover, galeri, denah, brosur

def brosur_path(slug):
    p = os.path.join(BROSUR_BASE, f"brosur-{slug}.pdf")
    return f"brosur/brosur-{slug}.pdf" if os.path.isfile(p) else None

def brosur_preview(slug):
    p = os.path.join(BROSUR_BASE, "preview", f"{slug}.png")
    return f"brosur/preview/{slug}.png" if os.path.isfile(p) else None

# ---------------- build ----------------
warnings = []
projects = []
for p in proj_rows:
    slug = p["slug_proyek"]
    cover, galeri, denah, brosur_img = scan_foto(slug)
    raw_tipe = tipe_by_slug.get(slug, [])
    info_x = INFO_TAMBAHAN.get(slug, {})

    tipe_unit = []
    harga_list, tipe_tersedia = [], 0
    for t in raw_tipe:
        harga = t.get("harga_rupiah")
        lt, lb = t.get("luas_tanah_m2"), t.get("luas_bangunan_m2")
        status_unit = t.get("status_unit") or "Tersedia"
        if isinstance(harga, (int, float)):
            harga_list.append(harga)
        if str(status_unit).strip().lower() == "tersedia":
            tipe_tersedia += 1
        # validasi janggal — hanya untuk residensial (komersial/ruko wajar LB tinggi krn 3-6 lantai)
        lantai = t.get("jumlah_lantai")
        if (p.get("kategori") == "Residensial"
                and isinstance(lt, (int, float)) and isinstance(lb, (int, float)) and lb
                and isinstance(lantai, (int, float)) and lantai
                and lt < (lb / lantai) * 0.5):
            warnings.append(f"{slug}/{t.get('nama_tipe')}: LT({lt}) janggal vs LB/lantai({lb}/{lantai}) — cek data")
        tipe_unit.append({
            "nama_tipe": t.get("nama_tipe"),
            "luas_tanah_m2": lt,
            "luas_bangunan_m2": lb,
            "kamar_tidur": t.get("kamar_tidur"),
            "kamar_mandi": t.get("kamar_mandi"),
            "jumlah_lantai": t.get("jumlah_lantai"),
            "harga_rupiah": harga,
            "harga_display": rupiah_singkat(harga),
            "status_unit": status_unit,
        })

    harga_min = min(harga_list) if harga_list else p.get("harga_mulai_dari")
    harga_max = max(harga_list) if harga_list else None

    # ringkasan untuk card (range dari tipe unit)
    def rng(key):
        vals = [t[key] for t in tipe_unit if isinstance(t.get(key), (int, float))]
        if not vals:
            return None
        lo, hi = min(vals), max(vals)
        # tampilkan bilangan bulat kalau bisa
        f = lambda x: int(x) if float(x).is_integer() else x
        return f"{f(lo)}" if lo == hi else f"{f(lo)}-{f(hi)}"

    teks_gabung = " ".join(str(p.get(k) or "") for k in
                           ("deskripsi_singkat", "keunggulan_utama", "fasilitas_kawasan"))

    projects.append({
        "slug": slug,
        "nama": p.get("nama_proyek"),
        "kategori": p.get("kategori"),
        "lokasi": p.get("lokasi_area"),
        "alamat": p.get("alamat_lengkap"),
        "status": p.get("status_proyek"),
        "harga_mulai_dari": harga_min,
        "harga_mulai_dari_display": rupiah_singkat(harga_min),
        "harga_tertinggi": harga_max,
        "harga_tertinggi_display": rupiah_singkat(harga_max),
        "deskripsi": p.get("deskripsi_singkat"),
        "keunggulan": p.get("keunggulan_utama"),
        "fasilitas": p.get("fasilitas_kawasan"),
        "fasilitas_tags": extract_tags(teks_gabung, FASILITAS_KEYWORDS),
        "poi_terdekat": extract_tags(teks_gabung, POI_KEYWORDS),
        "legalitas": p.get("legalitas"),
        "skema_pembayaran": p.get("skema_pembayaran"),
        # ---- info tambahan hasil bacaan brosur ----
        "highlight": info_x.get("highlight", []),
        "spesifikasi": info_x.get("spesifikasi", []),
        "fasilitas_detail": info_x.get("fasilitas_kawasan", []),
        "promo": info_x.get("promo", []),
        "arsitek": info_x.get("arsitek"),
        "link_video": p.get("link_video"),
        "wa_khusus": p.get("wa_khusus_proyek"),
        "urutan_tampil": p.get("urutan_tampil"),
        "featured": str(p.get("featured")).strip().upper() == "YA",
        "jumlah_tipe": len(tipe_unit),
        "jumlah_tipe_tersedia": tipe_tersedia,
        # ringkasan untuk card rekomendasi/listing
        "card": {
            "kamar_tidur": rng("kamar_tidur"),
            "kamar_mandi": rng("kamar_mandi"),
            "luas_tanah_m2": rng("luas_tanah_m2"),
            "luas_bangunan_m2": rng("luas_bangunan_m2"),
        },
        "foto": {"cover": cover, "galeri": galeri, "denah": denah,
                 "brosur": brosur_img,
                 "jumlah": (1 if cover else 0) + len(galeri) + len(denah) + len(brosur_img)},
        "brosur": brosur_path(slug),
        "brosur_preview": brosur_preview(slug),
        "tipe_unit": tipe_unit,
    })

projects.sort(key=lambda x: (not x["featured"], x["urutan_tampil"] or 9999))

# ---------------- rekomendasi: kategori sama + harga terdekat ----------------
def harga_key(p):
    return p["harga_mulai_dari"] if isinstance(p["harga_mulai_dari"], (int, float)) else 0

for p in projects:
    kandidat = [q for q in projects
                if q["slug"] != p["slug"] and q["kategori"] == p["kategori"]]
    # urutkan berdasarkan selisih harga_mulai_dari terkecil
    kandidat.sort(key=lambda q: abs(harga_key(q) - harga_key(p)))
    p["rekomendasi"] = [q["slug"] for q in kandidat[:6]]

# ---------------- statistik global ----------------
all_harga = [t["harga_rupiah"] for p in projects for t in p["tipe_unit"]
             if isinstance(t["harga_rupiah"], (int, float))]
by_kat, by_lok = {}, {}
for p in projects:
    by_kat[p["kategori"]] = by_kat.get(p["kategori"], 0) + 1
    by_lok[p["lokasi"]] = by_lok.get(p["lokasi"], 0) + 1

meta = {
    "sumber": "Paramount Land - Paramount Gading Serpong",
    "jumlah_proyek": len(projects),
    "jumlah_tipe_unit": sum(p["jumlah_tipe"] for p in projects),
    "jumlah_foto": sum(p["foto"]["jumlah"] for p in projects),
    "jumlah_brosur": sum(1 for p in projects if p["brosur"]),
    "harga_terendah": min(all_harga) if all_harga else None,
    "harga_terendah_display": rupiah_singkat(min(all_harga)) if all_harga else None,
    "harga_tertinggi": max(all_harga) if all_harga else None,
    "harga_tertinggi_display": rupiah_singkat(max(all_harga)) if all_harga else None,
    "proyek_per_kategori": by_kat,
    "proyek_per_lokasi": by_lok,
    "proyek_featured": [p["slug"] for p in projects if p["featured"]],
    "konteks_kota": KONTEKS_KOTA,
}

with open(OUT, "w", encoding="utf-8") as f:
    json.dump({"meta": meta, "proyek": projects}, f, ensure_ascii=False, indent=2)

print("OK ->", OUT)
print(json.dumps(meta, ensure_ascii=False, indent=2))
if warnings:
    print(f"\n=== {len(warnings)} WARNING VALIDASI ===")
    for w in warnings[:30]:
        print("  !", w)
